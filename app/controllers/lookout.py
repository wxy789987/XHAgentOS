# 瞭望管理控制层
import json
import logging
import re
import io
from datetime import datetime
import tornado.web
from concurrent.futures import ThreadPoolExecutor
from app.controllers.base import BaseHandler
from app.models.lookout import LookoutSourceRepository, LookoutRecordRepository

logger = logging.getLogger(__name__)
_executor = ThreadPoolExecutor(max_workers=4)


# ========== 瞭望源管理 ==========

class LookoutSourceHandler(BaseHandler):
    """瞭望源管理页面"""
    @tornado.web.authenticated
    def get(self):
        username = self.current_user if self.current_user else ""
        self.render("lookout_source.html", title="瞭望源管理", username=username)


class LookoutSourceListHandler(BaseHandler):
    """瞭望源列表API"""
    @tornado.web.authenticated
    def get(self):
        page = int(self.get_argument("page", 1))
        page_size = int(self.get_argument("page_size", 10))
        data = LookoutSourceRepository.get_page(page, page_size)
        self.write(data)


class LookoutSourceAddHandler(BaseHandler):
    """新增瞭望源"""
    @tornado.web.authenticated
    def post(self):
        name = self.get_argument("name", "")
        url_template = self.get_argument("url_template", "")
        headers_raw = self.get_argument("headers", "{}")
        item_selector = self.get_argument("item_selector", "")
        title_selector = self.get_argument("title_selector", "")
        url_selector = self.get_argument("url_selector", "")
        summary_selector = self.get_argument("summary_selector", "")
        time_selector = self.get_argument("time_selector", "")
        page_param = self.get_argument("page_param", "pn")
        page_size = int(self.get_argument("page_size", 10))

        if not name or not url_template:
            self.write({"code": 1, "msg": "名称和URL模板不能为空"})
            return

        try:
            headers = json.loads(headers_raw) if isinstance(headers_raw, str) else headers_raw
        except json.JSONDecodeError:
            headers = {}

        source_id = LookoutSourceRepository.create(
            name, url_template, headers, item_selector, title_selector,
            url_selector, summary_selector, time_selector, page_param, page_size,
        )
        self.write({"code": 0, "msg": "新增成功", "id": source_id})


class LookoutSourceEditHandler(BaseHandler):
    """编辑瞭望源"""
    @tornado.web.authenticated
    def post(self):
        source_id = int(self.get_argument("id", 0))
        name = self.get_argument("name", "")
        url_template = self.get_argument("url_template", "")
        headers_raw = self.get_argument("headers", "{}")
        item_selector = self.get_argument("item_selector", "")
        title_selector = self.get_argument("title_selector", "")
        url_selector = self.get_argument("url_selector", "")
        summary_selector = self.get_argument("summary_selector", "")
        time_selector = self.get_argument("time_selector", "")
        page_param = self.get_argument("page_param", "pn")
        page_size = int(self.get_argument("page_size", 10))
        is_active = int(self.get_argument("is_active", 1))

        if not source_id:
            self.write({"code": 1, "msg": "参数错误"})
            return

        try:
            headers = json.loads(headers_raw) if isinstance(headers_raw, str) else headers_raw
        except json.JSONDecodeError:
            headers = {}

        LookoutSourceRepository.update(
            source_id, name, url_template, headers, item_selector, title_selector,
            url_selector, summary_selector, time_selector, page_param, page_size, is_active,
        )
        self.write({"code": 0, "msg": "更新成功"})


class LookoutSourceGetHandler(BaseHandler):
    """获取单个瞭望源（含完整headers，供编辑回显）"""
    @tornado.web.authenticated
    def get(self):
        source_id = int(self.get_argument("id", 0))
        if not source_id:
            self.write({"code": 1, "msg": "参数错误"})
            return
        source = LookoutSourceRepository.get_by_id(source_id)
        if source:
            # 确保 headers 字段返回可解析的 JSON 字符串
            if source.get("headers"):
                try:
                    json.loads(source["headers"])
                except (json.JSONDecodeError, TypeError):
                    source["headers"] = "{}"
            else:
                source["headers"] = "{}"
            self.write({"code": 0, "data": source})
        else:
            self.write({"code": 1, "msg": "采集源不存在"})


class LookoutSourceDeleteHandler(BaseHandler):
    """删除瞭望源"""
    @tornado.web.authenticated
    def post(self):
        source_id = int(self.get_argument("id", 0))
        if source_id:
            LookoutSourceRepository.delete(source_id)
            self.write({"code": 0, "msg": "删除成功"})
        else:
            self.write({"code": 1, "msg": "参数错误"})


class LookoutSourceActiveHandler(BaseHandler):
    """获取启用的瞭望源列表"""
    @tornado.web.authenticated
    def get(self):
        sources = LookoutSourceRepository.get_active_sources()
        self.write({"code": 0, "items": sources})


# ========== 瞭望采集 ==========

class LookoutCollectPageHandler(BaseHandler):
    """瞭望采集界面"""
    @tornado.web.authenticated
    def get(self):
        username = self.current_user if self.current_user else ""
        self.render("lookout_collect.html", title="瞭望采集", username=username)


class LookoutCollectRunHandler(BaseHandler):
    """执行采集"""
    @tornado.web.authenticated
    async def post(self):
        import httpx
        from bs4 import BeautifulSoup

        source_id = int(self.get_argument("source_id", 0))
        keyword = self.get_argument("keyword", "")
        count = int(self.get_argument("count", 10))

        source = LookoutSourceRepository.get_by_id(source_id)
        if not source:
            self.write({"code": 1, "msg": "采集源不存在"})
            return

        def do_collect():
            try:
                url_template = source["url_template"]
                headers = json.loads(source["headers"]) if isinstance(source["headers"], str) else {}
                page_param = source["page_param"] or "pn"
                page_size = source["page_size"] or 10

                # 替换关键字
                url = url_template.replace("{keyword}", keyword or "")

                collected = []
                pages_needed = max(1, (count + page_size - 1) // page_size)

                for page_idx in range(pages_needed):
                    current_url = url
                    if page_idx > 0:
                        pn = page_idx * page_size
                        if "?" in current_url:
                            current_url += f"&{page_param}={pn}"
                        else:
                            current_url += f"?{page_param}={pn}"

                    resp = httpx.get(
                        current_url,
                        headers=headers if headers else {
                            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                        },
                        timeout=30.0,
                        follow_redirects=True,
                    )

                    if resp.status_code != 200:
                        continue

                    soup = BeautifulSoup(resp.text, "lxml")

                    # 通用解析策略：百度 news 结果
                    # 当前百度新闻使用 result-op c-container 类名
                    results = soup.select("#content_left > .result-op, #content_left > .c-container")
                    if not results:
                        results = soup.select("div[class*='result']:not([class*='result-molecule']), .c-container")
                    if not results:
                        results = soup.select("a[href*='baidu'][target='_blank']")
                        results = [r.parent for r in results if r.parent]

                    for item in results:
                        if len(collected) >= count:
                            break
                        title_el = item.select_one("h3 a, .c-title a, .t a")
                        if not title_el:
                            title_el = item.find('a', href=lambda x: x and x.startswith('http'))
                        if not title_el:
                            title_el = item.select_one("h3")

                        title = title_el.get_text(strip=True) if title_el else ""
                        link = title_el.get("href", "") if title_el else ""

                        # 过滤掉非新闻链接
                        if not (title and link and 'passport.baidu.com' not in link and 'chat.baidu.com' not in link):
                            continue

                        summary_el = item.select_one(".c-abstract, .c-summary, .summary, .content")
                        summary = summary_el.get_text(strip=True) if summary_el else ""

                        time_el = item.select_one(".c-color-gray, .c-time, .time, span[class*='time']")
                        publish_time = time_el.get_text(strip=True) if time_el else ""

                        collected.append({
                                "title": title,
                                "url": link,
                                "summary": summary,
                                "publish_time": publish_time,
                                "content": "",
                                "keywords": keyword,
                            })

                # 批量保存
                saved = LookoutRecordRepository.bulk_insert(source_id, collected)
                return {"code": 0, "msg": f"采集完成，新增 {saved} 条数据", "count": saved, "total": len(collected)}

            except Exception as e:
                logger.error(f"采集失败: {e}")
                return {"code": 1, "msg": f"采集失败: {str(e)}"}

        loop = tornado.ioloop.IOLoop.current()
        result = await loop.run_in_executor(_executor, do_collect)
        self.write(result)


# ========== 数据仓库 ==========

class LookoutWarehouseHandler(BaseHandler):
    """数据仓库页面"""
    @tornado.web.authenticated
    def get(self):
        username = self.current_user if self.current_user else ""
        self.render("lookout_warehouse.html", title="数据仓库", username=username)


class LookoutWarehouseListHandler(BaseHandler):
    """数据仓库列表API"""
    @tornado.web.authenticated
    def get(self):
        page = int(self.get_argument("page", 1))
        page_size = int(self.get_argument("page_size", 20))
        source_id = self.get_argument("source_id", None)
        keyword = self.get_argument("keyword", "")
        date_from = self.get_argument("date_from", "")
        date_to = self.get_argument("date_to", "")
        status = self.get_argument("status", "")
        deep_status = self.get_argument("deep_status", None)

        if source_id:
            source_id = int(source_id) if source_id else None

        data = LookoutRecordRepository.get_page(
            page=page, page_size=page_size,
            source_id=source_id, keyword=keyword,
            date_from=date_from, date_to=date_to, status=status,
            deep_status=deep_status,
        )
        self.write(data)


class LookoutWarehouseDeleteHandler(BaseHandler):
    """删除采集记录"""
    @tornado.web.authenticated
    def post(self):
        ids_raw = self.get_argument("ids", "")
        try:
            ids = json.loads(ids_raw) if ids_raw else []
        except json.JSONDecodeError:
            ids = []
        if ids:
            LookoutRecordRepository.delete(ids)
            self.write({"code": 0, "msg": f"已删除 {len(ids)} 条记录"})
        else:
            self.write({"code": 1, "msg": "请选择要删除的记录"})


# ========== 数据导出 ==========

class LookoutExportHandler(BaseHandler):
    """导出采集数据为Excel"""
    @tornado.web.authenticated
    def get(self):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        source_id = self.get_argument("source_id", None)
        keyword = self.get_argument("keyword", "")
        date_from = self.get_argument("date_from", "")
        date_to = self.get_argument("date_to", "")
        status = self.get_argument("status", "")

        if source_id:
            source_id = int(source_id) if source_id else None

        records = LookoutRecordRepository.get_all_for_export(
            source_id=source_id, keyword=keyword,
            date_from=date_from, date_to=date_to, status=status,
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "采集数据"

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        headers = ["ID", "采集源", "标题", "链接", "摘要", "发布时间", "关键词", "采集时间"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 数据行
        for i, rec in enumerate(records, 2):
            ws.cell(row=i, column=1, value=rec["id"]).border = thin_border
            ws.cell(row=i, column=2, value=rec.get("source_name", "")).border = thin_border
            ws.cell(row=i, column=3, value=rec["title"]).border = thin_border
            ws.cell(row=i, column=4, value=rec["url"]).border = thin_border
            ws.cell(row=i, column=5, value=rec.get("summary", "")).border = thin_border
            ws.cell(row=i, column=6, value=rec.get("publish_time", "")).border = thin_border
            ws.cell(row=i, column=7, value=rec.get("keywords", "")).border = thin_border
            ws.cell(row=i, column=8, value=rec.get("create_at", "")).border = thin_border

        # 列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 16
        ws.column_dimensions['H'].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        self.set_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.set_header("Content-Disposition", f"attachment; filename=lookout_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        self.write(output.read())

    def compute_etag(self):
        return None
